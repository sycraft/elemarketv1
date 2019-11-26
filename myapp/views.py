import os
import openpyxl
from django.shortcuts import render
from django.http import HttpResponse, FileResponse
from django.template import RequestContext
from myapp import models


def index(request):
    if request.method == "GET":

        wb2 = openpyxl.load_workbook('.\\resources\\psst_output.xlsx')
        excel_data1 = list()
        excel_data2 = list()
        excel_data3 = list()
        excel_data4 = list()
        excel_data5 = list()

        for i in models.Info1.objects.all():
            row_data = list()
            row_data.append(i.number)
            row_data.append(i.genname)
            row_data.append(i.bus)
            row_data.append(i.busname)
            row_data.append(i.Pmin)
            row_data.append(i.Pmax)
            row_data.append(i.minON)
            row_data.append(i.minOFF)
            row_data.append(i.Status)
            row_data.append(i.IntHour)
            row_data.append(i.IntPow)
            row_data.append(i.SUCost)
            row_data.append(i.SDCost)
            row_data.append(i.NOLOADCOST)
            row_data.append(i.k1)
            row_data.append(i.SP1)
            row_data.append(i.k2)
            row_data.append(i.SP2)
            row_data.append(i.k3)
            row_data.append(i.SP3)
            row_data.append(i.k4)
            row_data.append(i.SP4)
            row_data.append(i.RAMP_RATE)
            excel_data1.append(row_data)

        for i in models.Info2.objects.all():
            row_data = list()
            row_data.append(i.B1)
            row_data.append(i.B2)
            row_data.append(i.B3)
            row_data.append(i.B11)
            row_data.append(i.B12)
            row_data.append(i.B13)
            row_data.append(i.B14)
            row_data.append(i.B21)
            row_data.append(i.B22)
            row_data.append(i.B23)
            row_data.append(i.B24)
            row_data.append(i.B31)
            row_data.append(i.B32)
            row_data.append(i.B33)
            row_data.append(i.B101)
            row_data.append(i.B111)
            row_data.append(i.B231)
            row_data.append(i.B311)
            excel_data2.append(row_data)

        sheet_obj = wb2["obj"]
        for row in sheet_obj.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data3.append(row_data)

        sheet_P = wb2["P"]
        for row in sheet_P.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data4.append(row_data)

        sheet_LMP = wb2["LMP"]
        for row in sheet_LMP.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data5.append(row_data)

        return render(request, 'myapp/index.html', {"excel_data1": excel_data1,
                                                    "excel_data2": excel_data2,
                                                    "excel_data3": excel_data3,
                                                    "excel_data4": excel_data4,
                                                    "excel_data5": excel_data5, })
    else:
        excel_file = request.FILES["excel_file"]
        destination = open(os.path.join(".\\resources", 'psst_input.xlsx'), 'wb+')
        for chunk in excel_file.chunks():
            destination.write(chunk)
        destination.close()
        wb1 = openpyxl.load_workbook('.\\resources\\psst_input.xlsx')
        wb2 = openpyxl.load_workbook('.\\resources\\psst_output.xlsx')
        excel_data1 = list()
        excel_data2 = list()
        excel_data3 = list()
        excel_data4 = list()
        excel_data5 = list()

        sheet_gen = wb1["gen"]
        for row in sheet_gen.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data1.append(row_data)

        sheet_load = wb1["load"]
        for row in sheet_load.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data2.append(row_data)

        sheet_obj = wb2["obj"]
        for row in sheet_obj.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data3.append(row_data)

        sheet_P = wb2["P"]
        for row in sheet_P.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data4.append(row_data)

        sheet_LMP = wb2["LMP"]
        for row in sheet_LMP.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data5.append(row_data)

        return render(request, 'myapp/index.html', {"excel_data1": excel_data1,
                                                    "excel_data2": excel_data2,
                                                    "excel_data3": excel_data3,
                                                    "excel_data4": excel_data4,
                                                    "excel_data5": excel_data5, })


def download(request):
    file = open('.\\resources\\temp.xlsx', 'rb')
    response = HttpResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename="temp.xlsx'
    return response


def get_info(request):
    wb = openpyxl.load_workbook('.\\resources\\psst_input.xlsx')
    ws1, ws2 = wb["gen"], wb["load"]
    for r in range(1, 7):
        models.Info1.objects.create(number=str(ws1.cell(row=r, column=1).value),
                                    genname=str(ws1.cell(row=r, column=2).value),
                                    bus=str(ws1.cell(row=r, column=3).value),
                                    busname=str(ws1.cell(row=r, column=4).value),
                                    Pmin=str(ws1.cell(row=r, column=5).value),
                                    Pmax=str(ws1.cell(row=r, column=6).value),
                                    minON=str(ws1.cell(row=r, column=7).value),
                                    minOFF=str(ws1.cell(row=r, column=8).value),
                                    Status=str(ws1.cell(row=r, column=9).value),
                                    IntHour=str(ws1.cell(row=r, column=10).value),
                                    IntPow=str(ws1.cell(row=r, column=11).value),
                                    SUCost=str(ws1.cell(row=r, column=12).value),
                                    SDCost=str(ws1.cell(row=r, column=13).value),
                                    NOLOADCOST=str(ws1.cell(row=r, column=14).value),
                                    k1=str(ws1.cell(row=r, column=15).value),
                                    SP1=str(ws1.cell(row=r, column=16).value),
                                    k2=str(ws1.cell(row=r, column=17).value),
                                    SP2=str(ws1.cell(row=r, column=18).value),
                                    k3=str(ws1.cell(row=r, column=19).value),
                                    SP3=str(ws1.cell(row=r, column=20).value),
                                    k4=str(ws1.cell(row=r, column=21).value),
                                    SP4=str(ws1.cell(row=r, column=22).value),
                                    RAMP_RATE=str(ws1.cell(row=r, column=23).value))
    for r in range(1, 26):
        models.Info2.objects.create(B1=str(ws2.cell(row=r, column=2).value),
                                    B2=str(ws2.cell(row=r, column=3).value),
                                    B3=str(ws2.cell(row=r, column=4).value),
                                    B11=str(ws2.cell(row=r, column=5).value),
                                    B12=str(ws2.cell(row=r, column=6).value),
                                    B13=str(ws2.cell(row=r, column=7).value),
                                    B14=str(ws2.cell(row=r, column=8).value),
                                    B21=str(ws2.cell(row=r, column=9).value),
                                    B22=str(ws2.cell(row=r, column=10).value),
                                    B23=str(ws2.cell(row=r, column=11).value),
                                    B24=str(ws2.cell(row=r, column=12).value),
                                    B31=str(ws2.cell(row=r, column=13).value),
                                    B32=str(ws2.cell(row=r, column=14).value),
                                    B33=str(ws2.cell(row=r, column=15).value),
                                    B101=str(ws2.cell(row=r, column=16).value),
                                    B111=str(ws2.cell(row=r, column=17).value),
                                    B231=str(ws2.cell(row=r, column=18).value),
                                    B311=str(ws2.cell(row=r, column=19).value))
    return HttpResponse("Done")
